import random
from EnToBnValue import EnToBnValue
from openpyxl import Workbook, load_workbook


INS_NAME = "কাজী এন্টারপ্রাইজেস লিমিটেড"
INS_ADDRESS = "শোভাপুর, রাজফুলবাড়ীয়া, সাভার, ঢাকা"
PROD_NAME = "ডিশ ওয়াশিং বার"
PROD_DATE = "06-02-2023"
EXP_DATE = "06-02-2026"
PRICE = "546"
INSPECTION_DATE = "07-02-2023"


gross = [3.23,5,12,52.51,76.21,23.2,51.2]
empty_pack = [2.03,2.42,2.16]
avg = sum(gross)/len(gross)
for i in range(0,32-len(gross)):
    gross.append(round(random.uniform(avg*1.03,avg*0.97),2))
for i in range(0,32-len(empty_pack)):
    empty_pack.append(random.choice(empty_pack))

gross_in_bn = []
empty_pack_in_bn = []

for item in gross:
    val = EnToBnValue(item).in_bn_letter()
    gross_in_bn.append(val)

for item in empty_pack:
    val = EnToBnValue(item).in_bn_letter()
    empty_pack_in_bn.append(val)


wb = load_workbook("report.xlsx")
ws = wb.active

for i in range(13,45):
    ws[f'C{i}'].value = float(gross_in_bn[i-13])
for i in range(13,45):
    ws[f'D{i}'].value = float(empty_pack_in_bn[i-13])

for i in range(13,45):
    ws[f'E{i}'].value = f'=SUM(C{i}-D{i})'

ws["C45"] = "=AVERAGE(C13:C44)"
ws["D45"] = "=AVERAGE(D13:D44)"
ws["E45"] = "=SUM(C45-D45)"
wb.save("newReport.xlsx")
